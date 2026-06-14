#!/usr/bin/env python3
"""Export World Cup pool predictions from Excel to pool.json."""

from __future__ import annotations

import json
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "World Cup 2026.xlsx"
OUT = ROOT / "data" / "pool.json"


def serialize(value):
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def find_sheet(workbook, name: str) -> str | None:
    normalized = name.upper().replace(".", "").replace(" ", "").strip()
    for sheet_name in workbook.sheetnames:
        sheet_norm = sheet_name.upper().replace(".", "").replace(" ", "").strip()
        if sheet_norm == normalized:
            return sheet_name
    for sheet_name in workbook.sheetnames:
        if normalized.startswith(sheet_name.upper()[:4]) or sheet_name.upper().startswith(normalized[:4]):
            return sheet_name
    return None


def main() -> None:
    if not XLSX.exists():
        raise SystemExit(f"Missing Excel file: {XLSX}")

    workbook = openpyxl.load_workbook(XLSX, data_only=True)
    dashboard = workbook["Dashboard Final Score"]

    header = [dashboard.cell(4, col).value for col in range(1, dashboard.max_column + 1)]
    participants = [name.strip() if isinstance(name, str) else name for name in header[8:] if name]

    matches = []
    for row in range(6, dashboard.max_row + 1):
        match_id = dashboard.cell(row, 1).value
        if not isinstance(match_id, (int, float)):
            continue
        matches.append(
            {
                "id": int(match_id),
                "day": serialize(dashboard.cell(row, 2).value),
                "date": serialize(dashboard.cell(row, 3).value),
                "time": serialize(dashboard.cell(row, 4).value),
                "home": dashboard.cell(row, 5).value,
                "away": dashboard.cell(row, 8).value,
            }
        )

    predictions: dict[str, dict[int, dict[str, int | None]]] = {}
    for participant in participants:
        sheet_name = find_sheet(workbook, participant)
        if not sheet_name:
            raise SystemExit(f"No sheet found for participant: {participant}")

        sheet = workbook[sheet_name]
        participant_predictions: dict[int, dict[str, int | None]] = {}
        for row in range(6, sheet.max_row + 1):
            match_id = sheet.cell(row, 1).value
            if not isinstance(match_id, (int, float)):
                continue
            participant_predictions[int(match_id)] = {
                "home": sheet.cell(row, 6).value,
                "away": sheet.cell(row, 7).value,
            }
        predictions[participant] = participant_predictions

    OUT.parent.mkdir(exist_ok=True)
    payload = {
        "title": "Office World Cup 2026 Pool",
        "participants": participants,
        "matches": matches,
        "predictions": predictions,
        "scoring": {
            "exactScore": 1,
            "description": "1 point per exact score prediction",
        },
        "exportedAt": datetime.now().isoformat(),
    }

    OUT.write_text(json.dumps(payload, indent=2, default=str) + "\n", encoding="utf-8")
    print(f"Exported {len(participants)} participants and {len(matches)} matches to {OUT}")


if __name__ == "__main__":
    main()
